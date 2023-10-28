import imageio.v2 as imageio
import matplotlib.pyplot as plt
import numpy as np
from numpy import int32, array
from sklearn.cluster import KMeans
import cv2
import skimage
from skimage.color import rgb2lab, deltaE_cie76
from collections import Counter
import os 
import re
from openpyxl import Workbook
import colour
import time
import datetime

def YIndex(file_path, output_path, black, white,number):

    today = datetime.date.today()
    image=imageio.imread(str(file_path))

    subStr = file_path.split('/')[-1].split('.')[0]
    

    def RGB_HEX(color):
        return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2])) 

    def get_YI(image, number_of_colors, black, white):
        reshaped_image = cv2.resize(image, (600, 400))
        reshaped_image = reshaped_image.reshape(reshaped_image.shape[0]*reshaped_image.shape[1], 3)
        if white == True: 
            na = np.array([f for f in reshaped_image if f[0] !=255 and f[1] !=255 and f[2] !=255], np.uint8)
        elif black == True:
            na = np.array([f for f in reshaped_image if f[0] !=0 and f[1] !=0 and f[2] !=0], np.uint8)

        num_rows = na.shape
        num_rows = num_rows[0]
        clf = KMeans(n_clusters = number_of_colors, max_iter=300)
        labels = clf.fit_predict(na)
        counts = Counter(labels)
        counts = dict(sorted(counts.items()))
        center_colors = clf.cluster_centers_
        ordered_colors = [center_colors[i] for i in counts.keys()]
        hex_colors = [RGB_HEX(ordered_colors[i]) for i in counts.keys()]
        rgb_colors = [ordered_colors[int(i)] for i in counts.keys()]
        xyz= skimage.color.rgb2xyz([(rgb_colors[i]/255) for i in counts.keys()])
        xyz=xyz*100
        yellow = colour.colorimetry.yellowness_ASTME313([(xyz[i])for i in counts.keys()])

        plt.figure(figsize = (8, 6))
        plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)
        sample_file_name = '\ColorPalette '+ subStr + ' ' + str(today) 
        plt.savefig(str(output_path) + sample_file_name + ".jpg")

        return rgb_colors,xyz,hex_colors,yellow,num_rows,counts




    newwb = Workbook()
    newws = newwb.active
    newws.cell(row = 1, column = 1).value = "Unique_ID"
    newws.cell(row = 1, column = 2).value = "rgb_colors"
    newws.cell(row = 1, column = 3).value = "xyz"
    newws.cell(row = 1, column = 4).value = "HEX"
    newws.cell(row = 1, column = 5).value = "Yellowness Index (YI)"
    newws.cell(row = 1, column = 6).value = "Percentage"
    newws.cell(row = 1, column = 7).value = "YI proportion in Image"

    newws.cell(row = 2, column = 10).value = "Image's YI"

    Number_of_colours = int(number)

    rgb_colors,xyz,hex_colors,yellow,num_rows,counts = get_YI(image, Number_of_colours, black, white)


    vector = np.vectorize(np.int_)
    rgb_colors = vector(rgb_colors)
    xyz = vector(xyz)

    i=0
    x=2
    while i < Number_of_colours:
        print(i)
        newws.cell(row = x, column = 1).value = x-1
        newws.cell(row = x, column = 2).value = str(rgb_colors[i])
        newws.cell(row = x, column = 3).value = str(xyz[i])
        newws.cell(row = x, column = 4).value = str(hex_colors[i])
        if Number_of_colours == 1:
            newws.cell(row = x, column = 5).value = int(yellow)
        else:
            newws.cell(row = x, column = 5).value = round(yellow[i],2)
        newws.cell(row = x, column = 6).value = round(counts[i]/int(str(num_rows))*100,2)
        newws.cell(row = x, column = 7).value = round(newws.cell(row = x, column = 5).value*(newws.cell(row = x, column = 6).value/100),2)
        
        i=i+1
        x=x+1

    lastrow = newws.max_row
    col7_value=0

    for m in range(1, lastrow):
        x = newws.cell(row=m + 1, column=7).value
        col7_value=col7_value+x

    newws.cell(row = 3, column = 10).value = round(col7_value,1)
    print(col7_value)

    for column in newws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        newws.column_dimensions[column_letter].width = adjusted_width


    save_file = str(output_path) +'\YI analysis '+ subStr + ' ' + str(today) +'.xlsx'
    newwb.save(save_file)

#file_path = 'C:/Users/********************************************'
#output_path = R"C:\Users******************************************"
#black = True
#white = False
#number = str(5)

#YIndex(file_path, output_path, black, white, number)